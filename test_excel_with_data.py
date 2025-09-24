#!/usr/bin/env python3
"""
Test script to create an Excel template with sample data including the new boolean columns.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from pathlib import Path

# Headers from the reorganized template
HEADERS = [
    "Site Name or Business Name",
    "Person Contacted",
    "Tank Capacity",
    "Tank Measurements",
    "Has Dike",
    "Dike Measurements",
    "Diesel",
    "Pressurized Gas",
    "Acceptable Separation Distance Calculated",
    "Approximate Distance to Site (approximately)",
    "Compliance",
    "Additional information",
    "Latitude (NAD83)",
    "Longitude (NAD83)",
    "Calculated Distance to Polygon (ft)",
    "Tank Type",
]

# Sample data with different tank types (reorganized to match new column order)
SAMPLE_DATA = [
    ["ABC Gas Station", "John Smith", "10,000 gal", "12x8x10 ft", "Yes", "15x12x3 ft", "Yes", "No", "50 ft", "75 ft", "Yes", "Regular inspection", 18.234, -66.123, 75.5, "Underground"],
    ["XYZ Industrial", "Jane Doe", "5,000 gal", "8x6x8 ft", "Yes", "10x8x2 ft", "No", "Yes", "30 ft", "45 ft", "Yes", "Propane storage", 18.245, -66.134, 45.2, "Above Ground"],
    ["City Water Plant", "Bob Johnson", "20,000 gal", "20x10x12 ft", "Yes", "25x15x4 ft", "No", "No", "100 ft", "150 ft", "Yes", "Chemical storage", 18.256, -66.145, 150.0, "Above Ground"],
    ["Marine Fuel Depot", "Alice Brown", "15,000 gal", "15x10x10 ft", "Yes", "20x15x3 ft", "Yes", "No", "75 ft", "100 ft", "Yes", "Marine diesel", 18.267, -66.156, 100.3, "Above Ground"],
    ["Emergency Generator", "Tom Wilson", "500 gal", "4x3x4 ft", "No", "N/A", "Yes", "No", "25 ft", "30 ft", "Yes", "Backup power", 18.278, -66.167, 30.0, "Above Ground"],
]

def create_sample_excel():
    """Create an Excel file with sample data and formatting."""

    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Tank Data"

    # Add headers
    ws.append(HEADERS)

    # Format headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Add sample data
    for row_data in SAMPLE_DATA:
        ws.append(row_data)

    # Format boolean columns (Has Dike, Diesel, and Pressurized Gas)
    bool_fill_yes = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    bool_fill_no = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for row in range(2, ws.max_row + 1):
        # Has Dike column (5)
        dike_cell = ws.cell(row=row, column=5)
        if dike_cell.value == "Yes":
            dike_cell.fill = bool_fill_yes
        elif dike_cell.value == "No":
            dike_cell.fill = bool_fill_no

        # Diesel column (7)
        diesel_cell = ws.cell(row=row, column=7)
        if diesel_cell.value == "Yes":
            diesel_cell.fill = bool_fill_yes
        else:
            diesel_cell.fill = bool_fill_no

        # Pressurized Gas column (8)
        gas_cell = ws.cell(row=row, column=8)
        if gas_cell.value == "Yes":
            gas_cell.fill = bool_fill_yes
        else:
            gas_cell.fill = bool_fill_no

    # Adjust column widths
    column_widths = {
        'A': 25,  # Site Name
        'B': 15,  # Person Contacted
        'C': 12,  # Tank Capacity
        'D': 15,  # Tank Measurements
        'E': 10,  # Has Dike
        'F': 15,  # Dike Measurements
        'G': 10,  # Diesel
        'H': 15,  # Pressurized Gas
        'I': 20,  # Acceptable Separation
        'J': 20,  # Approximate Distance
        'K': 10,  # Compliance
        'L': 20,  # Additional info
        'M': 15,  # Latitude
        'N': 15,  # Longitude
        'O': 18,  # Calculated Distance
        'P': 15,  # Tank Type
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Freeze the header row
    ws.freeze_panes = "A2"

    # Save the file
    output_path = Path("output/excel_files/Tanks_With_Sample_Data.xlsx")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    print(f"✅ Excel file created with sample data!")
    print(f"   Path: {output_path.absolute()}")
    print(f"   Rows: {ws.max_row - 1} data rows")
    print(f"   Columns: {ws.max_column}")
    print("\n📊 Sample Data Summary:")
    print(f"   - Tanks with dike: {sum(1 for row in SAMPLE_DATA if row[4] == 'Yes')}")
    print(f"   - Tanks without dike: {sum(1 for row in SAMPLE_DATA if row[4] == 'No')}")
    print(f"   - Diesel tanks: {sum(1 for row in SAMPLE_DATA if row[6] == 'Yes')}")
    print(f"   - Pressurized gas tanks: {sum(1 for row in SAMPLE_DATA if row[7] == 'Yes')}")
    print(f"   - Other tanks: {sum(1 for row in SAMPLE_DATA if row[6] == 'No' and row[7] == 'No')}")

    wb.close()

if __name__ == "__main__":
    create_sample_excel()